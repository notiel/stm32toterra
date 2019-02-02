from dataclasses import dataclass, asdict
from typing import Dict, List, Union
import requests
from bs4 import BeautifulSoup, Tag
import sys
import datetime

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side

keys = ['Reference', 'Package', 'Flash', 'RAM']
base_url = r'https://www.terraelectronica.ru'
onelec_base = r'https://onelec.ru/products/'


@dataclass
class MicroController:
    partnumber: str      # code for cubemx
    is_available: bool   # is available in terraelectronica
    sales_data: list     # list of concrete pns, price, min quantity for order, instock, terra url, onelec url and price
    package: str         # mc package
    ram: int             # mc ram
    flash: int           # mc flash

    def __str__(self):
        return str(asdict(self))


def get_start_index(sheet)-> int:
    """
    gets index for title row
    :param sheet: sheet excel with data
    :return: index of table head row
    """
    for i in range(1, sheet.max_row):
        value = sheet['A%i' % i].value
        if value and 'Part No' in value:
            return i
    return 0


def get_column_indexes(sheet, i: int) -> Dict[str, str]:
    """
    gets indexes of service columns
    :param sheet: sheet of xlxs
    :param i: number of row with titles
    :return: dict with indices
    """
    indices = dict()
    for letter in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        value = sheet['%s%i' % (letter, i)].value
        if sheet['%s%i' % (letter, i)].value:
            for key in keys:
                if key in value:
                    indices[key] = letter
    return indices


def create_mc_list(sheet, indices: Dict[str, str], start_index: int) -> List[MicroController]:
    """
    create mc list from excel mk data
    :param sheet: excel shhet with data
    :param indices: list of icolumn indiced
    :param start_index: first row with data
    :return: list of microcontrollers
    """
    microcontrollers: List[MicroController] = list()
    for i in range(start_index+1, sheet.max_row+1):
        partnumber: str = sheet["%s%i" % (indices['Reference'], i)].value
        package: str = sheet["%s%i" % (indices['Package'], i)].value if 'Package' in indices.keys() else "Unknown"
        flash: str = sheet["%s%i" % (indices['Flash'], i)].value if 'Flash' in indices.keys() else "Unknown"
        ram: str = sheet["%s%i" % (indices['RAM'], i)].value if 'RAM' in indices.keys() else "Unknown"
        microcontroller = MicroController(partnumber=partnumber, is_available=False, sales_data=list(), package=package,
                                          flash=int(flash.split()[0]), ram=int(ram.split()[0]))
        microcontrollers.append(microcontroller)
    return microcontrollers


def get_delivery_info(pn: str, product_id: str) -> List[List[Union[float, int]]]:
    """
    function gets delivery data for product
    :param pn: product partnumber
    :param product_id: id of product
    :return: quantity available, number of delivery units, delivery unit: day or week, delivery prices
    """
    data = '{"jsonrpc":"2.0","method":"get_dms","params":{"code": %s, "name":"%s"}, "id": "objDMS||8"}'\
           % (product_id, pn)
    response = requests.post('https://www.terraelectronica.ru/services', data=data)
    res = response.text
    res = res.replace(r'\"', r'"')
    res = res.replace("\n", "")
    soup = BeautifulSoup(res)
    delivery_data = soup.find_all('tr')[1:]
    results = list()
    for item in delivery_data:
        try:
            price = float(item.contents[1].contents[1].contents[1].attrs['data-price'])
            count = int(item.contents[1].contents[1].contents[1].attrs['data-count'])
            instock = int(item.contents[3].contents[1].text)
            delivery = item.contents[5].contents[0]
            delivery = delivery.replace('\\', "")
            delivery = delivery.replace("n", "").strip()
            if "дн" in delivery:
                delivery = int(delivery.split()[0])
                if delivery < 14:
                    results.append([price, count, instock, delivery])
        except (ValueError, TypeError):
            print("Delivery exception")
    return results


def update_data_for_catalog(microcontroller: MicroController, request):
    """
    gets data for microcontroller for search redirected to catalog
    :param microcontroller: data for microcontroller to find
    :param request: request
    :return:
    """
    soup = BeautifulSoup(request.text)
    items = soup.find_all('tr')
    results = list()
    for item in items[1:]:
        pn, price, count, product_id, instock = "", 0, 0, "", 0
        try:

            content_url = item.contents[3]
            if content_url.attrs['class'][0] == 'table-item-name':
                content_url = content_url.contents[3]
                product_id = content_url.contents[0].attrs['href']
                pn = content_url.contents[0].contents[0]
            content_price = item.contents[11].contents[1].contents[1]
            if content_price.attrs['class'] == ['price-single', 'price-active']:
                price = float(content_price.attrs['data-price'])
                count = int(content_price.attrs['data-count'])
            content_count = item.contents[13].contents[1]
            if content_count.attrs['class'] == ['item-qnt']:
                instock = content_count.contents[0].split()[0]
            if int(instock) > 0:
                product_data = dict(PN=pn, Price=price, Count=count, Url=base_url + product_id, Instock=instock, Days=0)
                results.append(product_data)
            else:
                delivery_data = get_delivery_info(pn, product_id.split(r'/')[2])
                for offer in delivery_data:
                    product_data = dict(PN=pn, Price=offer[0], Count=offer[1], Url=base_url + product_id,
                                        Instock=offer[2], Days=offer[3])
                    results.append(product_data)

            onelec_results = get_onelec_data(pn)
            for result in onelec_results:
                product_data = dict(PN=pn, Price=result[0], Count=result[3], Url=result[4], Instock=result[2],
                                    Days=result[1])
                results.append(product_data)

        except (TypeError, IndexError, KeyError):
            print("error! in %s" % request.url)
            continue
    if len(results) > 0:
        microcontroller.sales_data = results.copy()
        microcontroller.is_available = True


def update_from_common_catalog(microcontroller: MicroController, request):
    """
     gets data from common catalog without filters
    :param microcontroller:
    :param request:
    :return:
    """
    soup = BeautifulSoup(request.text)
    items = soup.find_all('tr')
    results = list()
    for item in items[1:]:
        pn, price, count, product_id, instock = "", 0, 0, "", 0
        try:
            content_url = item.contents[3]
            if content_url.attrs['class'][0] == 'table-item-name':
                content_url = content_url.contents[3]
                product_id = content_url.contents[0].attrs['href']
                pn = content_url.contents[0].contents[0]
            content_price = item.contents[11].contents
            if len(content_price) > 1:
                if content_price[1].contents[1].attrs['class'] == ['price-single', 'price-active']:
                    price = float(content_price[1].contents[1].attrs['data-price'])
                    count = int(content_price[1].contents[1].attrs['data-count'])
                content_count = item.contents[13].contents[1]
                if content_count.attrs['class'] == ['item-qnt']:
                    instock = content_count.contents[0].split()[0]
            if int(instock) > 0:
                product_data = dict(PN=pn, Price=price, Count=count, Url=base_url + product_id, Instock=instock, Days=0)
                results.append(product_data)
            else:
                delivery_data = get_delivery_info(pn, product_id.split(r'/')[2])
                for offer in delivery_data:
                    product_data = dict(PN=pn, Price=offer[0], Count=offer[1], Url=base_url+product_id,
                                        Instock=offer[2], Days=offer[3])
                    results.append(product_data)
            onelec_offers = get_onelec_data(pn)
            for offer in onelec_offers:
                product_data = dict(PN=pn, Price=offer[0], Count=offer[3], Url=offer[4], Instock=offer[2], Days=offer[1])
                results.append(product_data)

        except (TypeError, IndexError, KeyError):
            print("error! in %s" % request.url)
            continue
    if len(results) > 0:
        microcontroller.sales_data = results.copy()
        microcontroller.is_available = True


def get_onelec_data(partnumber: str) -> List[List[Union[float, int, str]]]:
    """
    gets url and price from onelec
    :param partnumber: partnumber of product
    :return: price, url
    """
    url = onelec_base + partnumber.lower()
    res = requests.get(url)
    results = list()
    if res.status_code != 404:
        onelec_url = url
        soup = BeautifulSoup(res.text)
        table = soup.find('table', {'class': "table product-offers"})
        try:
            for tag in [tag for tag in table.contents[0].contents if isinstance(tag, Tag)]:
                try:
                    delivery = int(tag.contents[0].text.split()[1])
                    if delivery <= 14 and 'по запросу' not in tag.contents[1].text:
                        price = float(
                            tag.contents[2].contents[0].contents[0].attrs['data-price-rub'].split()[0].replace(',', '.'))
                        instock = int(tag.contents[0].contents[0].contents[4].text.split()[1])
                        min_order = tag.contents[2].contents[1].text.split()[1]
                        results.append([price, delivery, instock, min_order, onelec_url])
                except ValueError:
                    print("Onelec error at %s: " % onelec_url)
                    print("Data got: ")
                    print(results)
        except AttributeError:
            pass
    return results


def write_to_file(microcontrollers: List[MicroController]):
    """
    writes list of availale microcontrrollers to xlsx
    :param microcontrollers: list of microcontrollers
    :return:
    """
    wb = Workbook()
    wb.guess_types = True
    ws1 = wb.active
    ft = Font(bold=True)
    alignment = Alignment(horizontal='center', indent=0.2)
    border = Border(bottom=Side(border_style="double", color='00000000'))
    ws1.column_dimensions['A'].width = 20
    ws1.column_dimensions['B'].width = 20
    ws1.column_dimensions['F'].width = 50

    ws1['A1'] = "Model"
    ws1['B1'] = 'PN'
    ws1["C1"] = "Price"
    ws1['D1'] = 'Min count'
    ws1['E1'] = 'InStock'
    ws1['F1'] = 'Url'
    ws1['G1'] = 'Days'
    ws1['H1'] = 'Package'
    ws1['I1'] = 'Flash'
    ws1['J1'] = 'RAM'
    cells = [ws1['A1'], ws1['B1'], ws1['C1'], ws1['D1'], ws1['E1'], ws1['F1'], ws1['G1'], ws1['H1'], ws1['I1'],
             ws1['J1']]
    for cell in cells:
        cell.font = ft
        cell.border = border
        cell.alignment = alignment
    i = 2
    for microcontroller in microcontrollers:
        if microcontroller.is_available:
            for product in microcontroller.sales_data:
                ws1['A%i' % i] = microcontroller.partnumber
                ws1['B%i' % i] = product['PN']
                ws1['C%i' % i] = product['Price']
                ws1['D%i' % i] = product['Count']
                ws1['E%i' % i] = product['Instock']
                ws1['F%i' % i] = product['Url']
                ws1['G%i' % i] = product['Days']
                ws1['H%i' % i] = microcontroller.package
                ws1['I%i' % i] = microcontroller.flash
                ws1['J%i' % i] = microcontroller.ram
                i += 1
    now = datetime.datetime.now()
    wb.save(filename='Results'+(str(now).split('.')[0].replace(":", "-"))+'.xlsx')


def main(filename: str):

    try:
        wb = load_workbook(filename=filename)
        sheet = wb.active
    except FileNotFoundError:
        print("File %a not found" % filename)
        return

    start_index = get_start_index(sheet)
    if start_index == 0:
        print("No column titles")
        return

    indices = get_column_indexes(sheet, start_index)
    if 'Reference' not in indices.keys():
        print("No partnumber column, unable to proceed")
        return

    microcontrollers = create_mc_list(sheet, indices, start_index)

    for microcontroller in microcontrollers:
        search_text: str = microcontroller.partnumber
        search_text = search_text.replace('x', '')
        url: str = base_url + r"/search?text=" + search_text
        r = requests.get(url)
        print(r.url)
        if 'mikrokontrollery' in r.url:
            update_data_for_catalog(microcontroller, r)
        elif "search?" in r.url:
            soup = BeautifulSoup(r.text)
            links = soup.find('ul', {'class': "search-list"})
            if links:
                try:
                    link = base_url + '/' + links.contents[1].contents[1].attrs['href']
                    r = requests.get(link)
                    update_data_for_catalog(microcontroller, r)
                except (AttributeError, TypeError, IndexError):
                    print("Data error %s" % url)
        elif r'catalog/products/' in r.url:
            update_from_common_catalog(microcontroller, r)
        elif r'/product/' in r.url:
            pass
        else:
            print(r.url+' is not parced')

    write_to_file(microcontrollers)


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Please specify filename")
    else:
        main(sys.argv[1])
